"""XLSX text extraction via openpyxl.

Each worksheet becomes a ParsedPage. Cells are rendered as TSV rows so
the chunker / embedder can treat a sheet just like any other text and
retrieval still matches numeric / labeled content.

read_only=True streams the file row-by-row, keeping memory flat for
multi-MB workbooks. We skip empty leading/trailing whitespace rows.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from rag_service.parsers.types import ParsedDocument, ParsedPage


def _row_to_tsv(row: tuple) -> str:
    parts: list[str] = []
    for cell in row:
        if cell is None:
            parts.append("")
        else:
            # str() handles datetime / number / formula-result uniformly.
            value = str(cell).replace("\t", " ").replace("\n", " ")
            parts.append(value)
    return "\t".join(parts)


def parse_xlsx(content: bytes) -> ParsedDocument:
    """Extract every worksheet as a TSV-rendered ParsedPage."""
    workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    pages: list[ParsedPage] = []
    for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
        rows_text: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            line = _row_to_tsv(row).rstrip("\t")
            if line.strip():  # skip rows that are entirely empty
                rows_text.append(line)
        sheet_text = "\n".join(rows_text)
        pages.append(
            ParsedPage(
                page_number=sheet_index,
                text=sheet_text,
                extra={"sheet_name": sheet.title},
            )
        )
    workbook.close()

    full_text = "\n\n".join(
        f"### {p.extra['sheet_name']}\n{p.text}" for p in pages if p.text
    )
    return ParsedDocument(
        full_text=full_text,
        pages=pages,
        parser="openpyxl",
        extra={"sheet_count": len(pages)},
    )
